import MySQLdb
from Tkinter import *
import tkMessageBox
import tkSimpleDialog
import openpyxl
import  PIL
from PIL import ImageTk, Image
from openpyxl import Workbook
from openpyxl import load_workbook
from tkFileDialog import askopenfilename


db = MySQLdb.connect(host = "localhost", user = "root", passwd = "",  )
returnPath = "mysql"











##################################################################################################################################
############################################## ASSIGNMENT 2 FUNCTIONS ############################################################
##################################################################################################################################















def RepresentsInt(s):               #Checks whether int or not
    try: 
        int(s)
        return True
    except ValueError:
        return False












def logicalConverter(dList, rowcount,columncount,sSheet):
	
	i=0
	ws = wb[sSheet]
	while(i < rowcount-1):
		j=0
		while(j < columncount):
			
			dList.append(ws.cell(dList[0]).value)
			del dList[0]
			
			j = j + 1
		i=i+1


















def queryMakerVol2(aList,count,dbname) :
	print dbname
	exInsert = "INSERT INTO " + dbname + " values("
	
	while(count != 0):
		if(RepresentsInt(aList[0])):
			tempMem = aList[0]
			exInsert = exInsert + '%s'%tempMem + " , "
			count = count - 1
			del aList[0]
		else:
			tempMem = aList[0]
			exInsert = exInsert + "'%s'"%tempMem + " , "
			count = count - 1
			del aList[0]

	exInsert = exInsert[:-2] + ")"
	return exInsert


















def columnNameFetcher(bList):
	cursor = db.cursor()
	cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_schema = 'dbmn' AND table_name = 'emp' ")



	curs = db.cursor()
	curs.execute("SELECT Count(*) From information_schema.columns Where table_schema = 'dbmn' AND table_name = 'emp' " )
	a = curs.fetchall()
	curs.close()


	count = 0
	
	while count < a[0][0] :
		fy = cursor.fetchone()
		if(fy[0] in bList):          #If an element exists in db
			bList.remove(fy[0])
			count= count + 1
		else:
			cursor.close()
			tkMessageBox.showerror("Sorry","Element '%s' does not exist in SQL Table"%bList[0])
			return False

	cursor.close()
	return True




















def noOfSheets():

	strin = ""
	for sheet in wb:
		#print (sheet.title)
		strin = strin + sheet.title + "\n"


	tkMessageBox.showinfo("Sheets in '%s' ...          "%fname, strin)









def sheetToTable():

	sSheet = tkSimpleDialog.askstring("From '%s'"%fname,"Enter A Sheet name which exists in this excel file")
	dtabl = tkSimpleDialog.askstring("Table from '%s'"%des,"Enter A Table which exists in the Destination Database")

	if sSheet == "" or dtabl == "":
		tkMessageBox.showerror("Error","Please enter the proper information\n Terminating abrubtly!")
		exit()



	fdest = des + "." + dtabl

	maincursor = db.cursor()

	ws = wb[sSheet]
	if ws.cell(row=1, column = 1) == None:
		tkMessageBox.showerror("Error","Cannot Perform Because Sheet Is Empty.\n In Other words, there is nothing to Copy ")
	else :

		
		maincursor.execute("SELECT * FROM %s"%fdest)

		if maincursor.fetchone() == None:                                    #if the destination table selected does not contain any tuple/ rows of data


			aList = ['']

			for row in ws.rows:
				for cell in row:
					#print(cell.value)
					aList.append(cell.value)

			#print "\n"
			del aList[0]
			#print aList

			bList = ['']
			column_count = ws.max_column
			
			del bList[0]
			while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
				bList.append(aList[0])
				column_count = column_count - 1
				del aList[0]

			

			ero = ws.max_row

			#If bList is equal to table in database

			cur = db.cursor()
			s = columnNameFetcher(bList)
			
			
			if (s == True): #Meaning the columns in this sheet are same as tables (Upper description Columns)

				while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
					strr = queryMakerVol2(aList,ws.max_column,fdest)
					cur.execute(strr)
					print strr
					ero = ero - 1

			flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
			if flag == True:
				db.commit()	


		else:
			tempInput = tkSimpleDialog.askstring("Wait!","This table in the destination database allready consists of data. Do you wish to Overwrite? OR Do you wish to append (Press 'a' for append OR 'o' for overwrite)")


			if tempInput == 'a' or tempInput == 'A':
				
				aList = ['']

				for row in ws.rows:
					for cell in row:
						#print(cell.value)
						aList.append(cell.value)

				#print "\n"
				del aList[0]
				#print aList

				bList = ['']
				column_count = ws.max_column
				
				del bList[0]
				while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
					bList.append(aList[0])
					column_count = column_count - 1
					del aList[0]

				

				ero = ws.max_row

				#If bList is equal to table in database

				cur = db.cursor()
				s = columnNameFetcher(bList)
				
				
				if (s == True): #Meaning the columns in this sheet are same as tables (Upper description Columns)

					while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
						strr = queryMakerVol2(aList,ws.max_column,fdest)
						cur.execute(strr)
						print strr
						ero = ero - 1

				flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
				if flag == True:
					db.commit()	

			elif tempInput == 'o' or tempInput == 'O':
				
				maincursor.execute("TRUNCATE TABLE %s"%fdest)							#Overwrite Starts here
				aList = ['']

				for row in ws.rows:
					for cell in row:
						#print(cell.value)
						aList.append(cell.value)

				#print "\n"
				del aList[0]
				#print aList

				bList = ['']
				column_count = ws.max_column
				
				del bList[0]
				while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
					bList.append(aList[0])
					column_count = column_count - 1
					del aList[0]

				

				ero = ws.max_row

				#If bList is equal to table in database

				cur = db.cursor()
				s = columnNameFetcher(bList)
				
				
				if (s == True): #Meaning the columns in this sheet are same as tables (Upper description Columns)

					while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
						strr = queryMakerVol2(aList,ws.max_column,fdest)
						cur.execute(strr)
						print strr
						ero = ero - 1

				flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
				if flag == True:
					db.commit()	
																				#Overwrite Ends Here

				else:
					tkMessageBox.showerror("Error","Cannot Perform Overwrite Because Both tables have different Number of Columns")#Append Ends here



			else:
				print "Sorry wrong option Entered" 
				exit()



		











# row_count = ws.max_row
	# column_count = ws.max_column
	# print row_count * column_count

def c2c():
	sSheet = tkSimpleDialog.askstring("From '%s'"%fname,"Enter A Sheet name which exists in this excel file")
	dtabl = tkSimpleDialog.askstring("Table from '%s'"%des,"Enter A Table which exists in the Destination Database")

	if sSheet == "" or dtabl == "":
		tkMessageBox.showerror("Error","Please enter the proper information\n Terminating abrubtly!")
		exit()



	fdest = des + "." + dtabl

	maincursor = db.cursor()

	ws = wb[sSheet]
	if ws.cell(row=1, column = 1) == None:
		tkMessageBox.showerror("Error","Cannot Perform Because Sheet Is Empty.\n In Other words, there is nothing to Copy ")
	else :

		maincursor.execute("SELECT * FROM %s"%fdest)

		if maincursor.fetchone() == None:


			c2cro = Tk()
			cList= ['']
			del cList[0]

			row_count = ws.max_row
			while(row_count-1 != 0):
				slm = tkSimpleDialog.askstring("From '%s'"%fname,"Enter A Column to map to '%s'"%dtabl)
				cList.append(slm)
				row_count = row_count - 1

			print cList

			colR = ws.max_row
			colC = ws.max_column
			
			i=1
			j=1
			dList = ['']

			i = 1
			count = 0
			while(i <= colR):                                           # THIS WHILE LOOP BASICALLY CREATES THE COLUMNS IN ORDER OF WHICH WE HAVE TO MAP
	   			j = 1
	   			while(j <= colC):
	   				#print str(cList[count] + str(i))
	   				dList.append(str(cList[count] + str(i)))

	   				count = count + 1
	   				j = j + 1

	   			i = i + 1
	   			count = 0

					
		   	del dList[0]
		   	#print ws.cell(dList[0]).value
		   	#print dList           #./
		   	column_count = ws.max_column
			
			cList= ['']	
			del cList[0]
			while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
				cList.append(dList[0])
				column_count = column_count - 1
				del dList[0]

			

			cur = db.cursor()

			ero = ws.max_row

			logicalConverter(dList,ws.max_row,ws.max_column,sSheet)

			

			while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
				strr = queryMakerVol2(dList,ws.max_column,fdest)
				cur.execute(strr)
				print strr
				ero = ero - 1

			flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
			if flag == True:
				db.commit()	
			cur.close()


			c2cro = mainloop()



		else:
			tempInput = tkSimpleDialog.askstring("Wait!","This table in the destination database allready consists of data. Do you wish to Overwrite? OR Do you wish to append (Press 'a' for append OR 'o' for overwrite)")


			if tempInput == 'a' or tempInput == 'A':
				
				c2cro = Tk()
				cList= ['']
				del cList[0]

				row_count = ws.max_row
				while(row_count-1 != 0):
					slm = tkSimpleDialog.askstring("From '%s'"%fname,"Enter A Column to map to '%s'"%dtabl)
					cList.append(slm)
					row_count = row_count - 1

				print cList

				colR = ws.max_row
				colC = ws.max_column
				
				i=1
				j=1
				dList = ['']

				i = 1
				count = 0
				while(i <= colR):                                           # THIS WHILE LOOP BASICALLY CREATES THE COLUMNS IN ORDER OF WHICH WE HAVE TO MAP
		   			j = 1
		   			while(j <= colC):
		   				#print str(cList[count] + str(i))
		   				dList.append(str(cList[count] + str(i)))

		   				count = count + 1
		   				j = j + 1

		   			i = i + 1
		   			count = 0

						
			   	del dList[0]
			   	#print ws.cell(dList[0]).value
			   	#print dList           #./
			   	column_count = ws.max_column
				
				cList= ['']	
				del cList[0]
				while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
					cList.append(dList[0])
					column_count = column_count - 1
					del dList[0]

				

				cur = db.cursor()

				ero = ws.max_row

				logicalConverter(dList,ws.max_row,ws.max_column,sSheet)

				

				while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
					strr = queryMakerVol2(dList,ws.max_column,fdest)
					cur.execute(strr)
					print strr
					ero = ero - 1

				flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
				if flag == True:
					db.commit()	
				cur.close()


				c2cro = mainloop()















			elif tempInput == 'o' or tempInput == 'O':
				
				maincursor.execute("TRUNCATE TABLE %s"%fdest)							#Overwrite Starts here
				c2cro = Tk()
				cList= ['']
				del cList[0]

				row_count = ws.max_row
				while(row_count-1 != 0):
					slm = tkSimpleDialog.askstring("From '%s'"%fname,"Enter A Column to map to '%s'"%dtabl)
					cList.append(slm)
					row_count = row_count - 1

				print cList

				colR = ws.max_row
				colC = ws.max_column
				
				i=1
				j=1
				dList = ['']

				i = 1
				count = 0
				while(i <= colR):                                           # THIS WHILE LOOP BASICALLY CREATES THE COLUMNS IN ORDER OF WHICH WE HAVE TO MAP
		   			j = 1
		   			while(j <= colC):
		   				#print str(cList[count] + str(i))
		   				dList.append(str(cList[count] + str(i)))

		   				count = count + 1
		   				j = j + 1

		   			i = i + 1
		   			count = 0

						
			   	del dList[0]
			   	#print ws.cell(dList[0]).value
			   	#print dList           #./
			   	column_count = ws.max_column
				
				cList= ['']	
				del cList[0]
				while(column_count != 0):                 #ONLY  FOR THE COLUMNFETCHER METHOD
					cList.append(dList[0])
					column_count = column_count - 1
					del dList[0]

				

				cur = db.cursor()

				ero = ws.max_row

				logicalConverter(dList,ws.max_row,ws.max_column,sSheet)

				

				while(ero -1 != 0):               #Number of rows - 1 because i have allready removed the column description row , this generated the queries needed
					strr = queryMakerVol2(dList,ws.max_column,fdest)
					cur.execute(strr)
					print strr
					ero = ero - 1

				flag = tkMessageBox.askyesno("Copy Machine B)","Tables have successfully been copied, Do you wish to COMMIT?")
				if flag == True:
					db.commit()	
				cur.close()


				c2cro = mainloop()



			else:
				print "Sorry wrong option Entered" 
				exit()

















##################################################################################################################################
############################################## ASSIGNMENT 2 FUNCTIONS END ########################################################
##################################################################################################################################












def donothingg():
	tkMessageBox.showinfo("Information","Application Created By Muhammad Shahriyar Sheikh \n Roll Number: 13K-2015")

def donothing():
	root.destroy()
	exit()


def databases():
	tab = db.cursor()
	tab.execute("Show Databases")
	condition = True
	stri = ""
	while condition == True:
		re = tab.fetchone()
		stri = stri + str(re) + '\n'
		if re == None:
			condition = False

		
	tkMessageBox.showinfo("Available Databases ...          ",stri[:-5])
	tab.close()



def tablesindest():
	tab = db.cursor()
	tab.execute("Show tables from %s"%des)
	re = tab.fetchall()
	tkMessageBox.showinfo("Tables in '%s' ...          "%des, re)
	tab.close()





if __name__ == "__main__":


	root = Tk()
	root.withdraw()           #make window dissappear
	databases()



	wb = Workbook()
	ws = wb.active
	fname = askopenfilename(filetypes=(("Excel Files", "*.xlsx"),
                                           ("HTML files", "*.html;*.htm"),
                                           ("All files", "*.*") ))
	wb = load_workbook(fname)







	des = tkSimpleDialog.askstring("Destination Database", "Enter the Destination Database (Where do you want to copy to)")
	cur2 = db.cursor()                          #To Check the existence of the second db
	secDbExists = cur2.execute("Show Databases like '%s'"%des)
	if secDbExists != 1:
		tkMessageBox.showwarning("Nope","Cannot Find any database named '%s'" % sou)
		exit()

	cur2.close()

	root.deiconify()






	path = "cppl.png"
	imga = ImageTk.PhotoImage(Image.open(path))
	panel = Label(root, image = imga)
	panel.pack(side = "top", fill = "both")




	menu = Menu(root)
	root.config(menu = menu)          #ima gonna cinfigure a parameter

	submenu = Menu(menu)                #menu item within a menu item
	menu.add_cascade(label = "File", menu = submenu)
	submenu.add_command(label = "About The Creator", command = donothingg)
	submenu.add_command(label = "Exit", command = donothing)








	topFrame = Frame(root)  #imma gonna make an invisible window which is going to take root
	topFrame.pack()         #to display anything, we need to pack it in
	bottomFrame = Frame(root)
	bottomFrame.pack(side = BOTTOM)      # parameter tells us were to put





	b1 = Button(topFrame,text = "Press This Button to view Sheet name inside the Excel File", command = noOfSheets)
	b2 = Button(topFrame,text = "Press This Button to view all tables inside '%s' Database"%des, command = tablesindest)
	b3 = Button(bottomFrame,text = "Copy From Sheet To Table", command = sheetToTable)
	b4 = Button(bottomFrame,text = "Copy From Column to column", command = c2c)








	b1.pack(side = LEFT)
	b1.config(font=('helvetica', 10), bd=8, relief=RAISED)
	b2.pack(side = LEFT)
	b2.config(font=('helvetica', 10), bd=8, relief=RAISED)





	path = "Excel-icon.png"
	#Creates a Tkinter-compatible photo image, which can be used everywhere Tkinter expects an image object.
	img = ImageTk.PhotoImage(Image.open(path))
	#The Label widget is a standard Tkinter widget used to display a text or image on the screen.
	panel = Label(root, image = img)
	#The Pack geometry manager packs widgets in rows or columns.
	panel.pack(side = "bottom", fill = "both", expand = "yes")





	b3.pack(side=LEFT)
	b3.config(font=('helvetica', 10), bd=8, relief=RAISED)

	b4.pack(side=BOTTOM)
	b4.config(font=('helvetica', 10), bd=8, relief=RAISED)



	root = mainloop()



